from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parent
INPUT_PATH = ROOT / "futbrain_enrichment.json"
OUTPUT_PATH = ROOT / "futbrain_db_enriched_v3.xlsx"


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def auto_fit(ws) -> None:
    for column_cells in ws.columns:
        width = 0
        letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, len(value))
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 70)


def clubs_compact(clubs: list[dict]) -> str:
    return " | ".join(
        f"{club['club_name']} ({club.get('from_year') or '?'}-{club.get('to_year') or '?'})"
        for club in clubs
    )


def main() -> None:
    data = json.loads(INPUT_PATH.read_text(encoding="utf-8"))
    players = data["players"]

    wb = Workbook()
    ws_players = wb.active
    ws_players.title = "Jugadores"
    ws_players.append(
        [
            "player_id",
            "name",
            "flag",
            "wiki_title",
            "wikipedia_url",
            "wikidata_url",
            "photo_url",
            "birth_year",
            "country",
            "current_db_clubs",
            "wikidata_clubs",
            "status",
            "notes",
        ]
    )

    ws_player_clubs = wb.create_sheet("Jugador_Clubes")
    ws_player_clubs.append(
        [
            "player_id",
            "player_name",
            "club_name",
            "from_year",
            "to_year",
            "source",
            "verification",
            "notes",
        ]
    )

    club_to_players: dict[str, set[str]] = defaultdict(set)
    review_rows: list[list] = []

    for player in players:
        current_comp = clubs_compact(player.get("current_clubs", []))
        wikidata_comp = clubs_compact(player.get("wikidata_clubs", []))

        ws_players.append(
            [
                player["player_id"],
                player["name"],
                player.get("flag", ""),
                player.get("query_name", player["name"]),
                player.get("wikipedia_url", ""),
                player.get("wikidata_url", ""),
                player.get("photo_url", ""),
                player.get("birth_year"),
                player.get("country", ""),
                current_comp,
                wikidata_comp,
                player.get("status", ""),
                player.get("notes", ""),
            ]
        )

        for club in player.get("wikidata_clubs", []):
            club_to_players[club["club_name"]].add(player["name"])
            ws_player_clubs.append(
                [
                    player["player_id"],
                    player["name"],
                    club["club_name"],
                    club.get("from_year"),
                    club.get("to_year"),
                    "Wikidata",
                    "",
                    "",
                ]
            )

        if player.get("status") != "ok":
            review_rows.append(
                [
                    player["player_id"],
                    player["name"],
                    player.get("status", ""),
                    current_comp,
                    wikidata_comp,
                    player.get("wikipedia_url", ""),
                    player.get("notes", ""),
                ]
            )

    style_header(ws_players)
    auto_fit(ws_players)
    style_header(ws_player_clubs)
    auto_fit(ws_player_clubs)

    ws_clubs = wb.create_sheet("Clubes")
    ws_clubs.append(
        ["club_id", "club_name", "player_count", "players", "country", "league", "badge_url", "notes"]
    )
    for idx, club_name in enumerate(sorted(club_to_players.keys()), start=1):
        players_for_club = sorted(club_to_players[club_name])
        ws_clubs.append(
            [
                f"C{idx:04d}",
                club_name,
                len(players_for_club),
                " | ".join(players_for_club),
                "",
                "",
                "",
                "",
            ]
        )
    style_header(ws_clubs)
    auto_fit(ws_clubs)

    ws_coaches = wb.create_sheet("Entrenadores")
    ws_coaches.append(["coach_id", "name", "country", "photo_url", "teams_coached", "notes"])
    style_header(ws_coaches)
    auto_fit(ws_coaches)

    ws_titles = wb.create_sheet("Titulos")
    ws_titles.append(
        [
            "title_id",
            "player_id",
            "player_name",
            "title_name",
            "title_type",
            "club_name",
            "year",
            "shared_year_label",
            "country",
            "source",
            "verified",
            "notes",
        ]
    )
    style_header(ws_titles)
    auto_fit(ws_titles)

    ws_titles_review = wb.create_sheet("Titulos_Revisar")
    ws_titles_review.append(["player_id", "player_name", "priority", "status", "notes"])
    for player in players:
        ws_titles_review.append(
            [
                player["player_id"],
                player["name"],
                "alta",
                "pendiente",
                "Completar titulos oficiales por club y ano.",
            ]
        )
    style_header(ws_titles_review)
    auto_fit(ws_titles_review)

    ws_wikipedia = wb.create_sheet("Wikipedia_Revisar")
    ws_wikipedia.append(
        [
            "player_id",
            "player_name",
            "wiki_title",
            "wikipedia_url",
            "wikidata_url",
            "photo_url",
            "status",
            "birth_year",
            "country",
            "current_db_clubs",
            "wikidata_clubs",
            "candidate_count",
            "review_priority",
            "review_action",
            "notes",
        ]
    )
    for player in players:
        status = player.get("status", "")
        candidate_count = len(player.get("candidates", []))
        has_photo = bool(player.get("photo_url"))
        review_priority = "alta" if status != "ok" or not has_photo else "media"

        ws_wikipedia.append(
            [
                player["player_id"],
                player["name"],
                player.get("query_name", player["name"]),
                player.get("wikipedia_url", ""),
                player.get("wikidata_url", ""),
                player.get("photo_url", ""),
                status,
                player.get("birth_year"),
                player.get("country", ""),
                clubs_compact(player.get("current_clubs", [])),
                clubs_compact(player.get("wikidata_clubs", [])),
                candidate_count,
                review_priority,
                "Verificar Wikipedia y completar titulos",
                player.get("notes", ""),
            ]
        )
    style_header(ws_wikipedia)
    auto_fit(ws_wikipedia)

    ws_review = wb.create_sheet("Revisar")
    ws_review.append(
        ["player_id", "name", "status", "current_db_clubs", "wikidata_clubs", "wikipedia_url", "notes"]
    )
    for row in review_rows:
        ws_review.append(row)
    style_header(ws_review)
    auto_fit(ws_review)

    ws_summary = wb.create_sheet("Resumen")
    ws_summary.append(["Metrica", "Valor"])
    ws_summary.append(["Jugadores cargados", len(players)])
    ws_summary.append(["Clubes unicos desde Wikidata", len(club_to_players)])
    ws_summary.append(["Jugadores OK", sum(1 for player in players if player.get("status") == "ok")])
    ws_summary.append(["Jugadores para revisar", len(review_rows)])
    ws_summary.append(["Jugadores sin foto_url", '=COUNTBLANK(Jugadores!G2:G9999)'])
    ws_summary.append(["Jugadores sin Wikidata", '=COUNTIF(Jugadores!L2:L9999,"sin_wikidata")'])
    ws_summary.append(["Jugadores sin clubes", '=COUNTIF(Jugadores!L2:L9999,"sin_clubes")'])
    ws_summary.append(["Titulos cargados", '=COUNTA(Titulos!A2:A99999)'])
    ws_summary.append(["Jugadores con titulos pendientes", '=COUNTIF(Titulos_Revisar!D2:D9999,"pendiente")'])
    ws_summary.append(["Filas para revisar en Wikipedia", '=COUNTA(Wikipedia_Revisar!A2:A9999)'])
    ws_summary.append(["Generado", data.get("generated_at", "")])
    ws_summary.append(["Nota", "Los casos ambiguos quedaron en la hoja Revisar."])
    style_header(ws_summary)
    auto_fit(ws_summary)

    wb.save(OUTPUT_PATH)
    print(f"Excel enriquecido: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
