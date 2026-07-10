from __future__ import annotations

import argparse
import json
import sys
import time
import unicodedata
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_INPUT = ROOT / "futbrain_db_con_titulos.xlsx"
DEFAULT_OUTPUT = ROOT / "futbrain_db_con_titulos_api_football.xlsx"
API_BASE = "https://v3.football.api-sports.io"


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    for old, new in {
        "'": " ",
        ".": " ",
        "-": " ",
        "_": " ",
        "&": " and ",
        "fc": " ",
        "cf": " ",
        "club": " ",
        "football": " ",
        "futbol": " ",
        "futebol": " ",
    }.items():
        text = text.replace(old, new)
    return " ".join(text.split())


def compact_clubs(clubs: list[tuple[str, int | None, int | None]]) -> str:
    return " | ".join(f"{club} ({frm or '?'}-{to or '?'})" for club, frm, to in clubs)


def parse_year(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    if len(text) >= 4 and text[:4].isdigit():
        return int(text[:4])
    return None


def copy_sheet_values(src_ws, dst_ws) -> None:
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
            if cell.number_format:
                new_cell.number_format = cell.number_format
            if cell.font:
                new_cell.font = copy(cell.font)
            if cell.fill:
                new_cell.fill = copy(cell.fill)
            if cell.border:
                new_cell.border = copy(cell.border)
            if cell.alignment:
                new_cell.alignment = copy(cell.alignment)
            if cell.protection:
                new_cell.protection = copy(cell.protection)
    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key] = copy(dim)
    for key, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[key] = copy(dim)
    dst_ws.freeze_panes = src_ws.freeze_panes
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))


@dataclass
class ApiFootballClient:
    api_key: str
    pause_seconds: float = 0.2

    def get(self, endpoint: str, **params: Any) -> dict[str, Any]:
        query = urlencode({k: v for k, v in params.items() if v not in (None, "")})
        url = f"{API_BASE}/{endpoint}"
        if query:
            url = f"{url}?{query}"
        request = Request(
            url,
            headers={
                "x-apisports-key": self.api_key,
                "Accept": "application/json",
                "User-Agent": "FutBrain-DB-Builder/1.0",
            },
        )
        try:
            with urlopen(request, timeout=30) as response:
                data = json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"API error {exc.code} on {endpoint}: {body}") from exc
        except URLError as exc:
            raise RuntimeError(f"Network error on {endpoint}: {exc}") from exc
        time.sleep(self.pause_seconds)
        return data


class FutbrainApiEnricher:
    def __init__(self, workbook_path: Path, output_path: Path, api_key: str) -> None:
        self.workbook_path = workbook_path
        self.output_path = output_path
        self.client = ApiFootballClient(api_key=api_key)
        self.wb = load_workbook(workbook_path)
        self.players_ws = self.wb["Jugadores"]
        self.clubs_ws = self.wb["Jugador_Clubes"]
        self.titles_ws = self.wb["Titulos"]
        self.review_ws = self.wb["Revisar"] if "Revisar" in self.wb.sheetnames else self.wb.create_sheet("Revisar")
        self.api_review_ws = self.wb["API_Football_Revisar"] if "API_Football_Revisar" in self.wb.sheetnames else self.wb.create_sheet("API_Football_Revisar")
        self.player_headers = self._header_map(self.players_ws)
        self.club_headers = self._header_map(self.clubs_ws)
        self.title_headers = self._header_map(self.titles_ws)
        self.summary: Counter[str] = Counter()
        self.transfer_rows_by_player: defaultdict[str, list[int]] = defaultdict(list)
        for row in range(2, self.clubs_ws.max_row + 1):
            pid = self.clubs_ws.cell(row=row, column=self.club_headers["player_id"]).value
            if pid:
                self.transfer_rows_by_player[str(pid)].append(row)

    @staticmethod
    def _header_map(ws) -> dict[str, int]:
        return {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

    def set_cell(self, ws, row: int, key: str, value: Any, headers: dict[str, int]) -> None:
        if key in headers:
            ws.cell(row=row, column=headers[key]).value = value

    def find_best_player_match(self, name: str) -> tuple[dict[str, Any] | None, str]:
        candidates: list[dict[str, Any]] = []
        attempted_endpoints = [
            ("players/profiles", {"search": name}),
            ("players", {"search": name}),
        ]
        last_error = ""
        for endpoint, params in attempted_endpoints:
            try:
                data = self.client.get(endpoint, **params)
                response = data.get("response") or []
                if response:
                    candidates = response
                    break
            except RuntimeError as exc:
                last_error = str(exc)
        if not candidates:
            return None, last_error or "Sin respuesta de players search"

        target = normalize_text(name)

        def score(item: dict[str, Any]) -> tuple[int, int]:
            player = item.get("player", item)
            first = normalize_text(player.get("firstname") or "")
            last = normalize_text(player.get("lastname") or "")
            full = normalize_text(player.get("name") or f"{first} {last}")
            exact = int(full == target)
            contains = int(target in full or full in target)
            return (exact, contains)

        ordered = sorted(candidates, key=score, reverse=True)
        best = ordered[0]
        return best, ""

    def extract_player_profile(self, match: dict[str, Any]) -> dict[str, Any]:
        player = match.get("player", match)
        birth = player.get("birth") or {}
        nationality = player.get("nationality")
        return {
            "api_player_id": player.get("id"),
            "full_name": player.get("name"),
            "birth_year": parse_year(birth.get("date")),
            "country": nationality,
            "photo_url": player.get("photo"),
        }

    def fetch_transfers(self, player_id: int) -> list[tuple[str, int | None, int | None]]:
        endpoints = [
            ("transfers", {"player": player_id}),
            ("players/transfers", {"player": player_id}),
        ]
        response: list[dict[str, Any]] = []
        for endpoint, params in endpoints:
            try:
                data = self.client.get(endpoint, **params)
                response = data.get("response") or []
                if response:
                    break
            except RuntimeError:
                continue

        clubs: list[tuple[str, int | None, int | None]] = []
        for item in response:
            transfers = item.get("transfers") if isinstance(item, dict) and "transfers" in item else response
            if transfers is response:
                pass
            for transfer in transfers:
                team_in = transfer.get("teams", {}).get("in") or {}
                joined = parse_year(transfer.get("date"))
                club_name = team_in.get("name")
                if club_name:
                    clubs.append((club_name, joined, None))
            break

        deduped: list[tuple[str, int | None, int | None]] = []
        seen: set[tuple[str, int | None]] = set()
        for club_name, frm, to in clubs:
            key = (club_name, frm)
            if key not in seen:
                seen.add(key)
                deduped.append((club_name, frm, to))
        return deduped

    def fetch_trophies(self, player_id: int) -> list[dict[str, Any]]:
        endpoints = [
            ("trophies", {"player": player_id}),
            ("players/trophies", {"player": player_id}),
        ]
        response: list[dict[str, Any]] = []
        for endpoint, params in endpoints:
            try:
                data = self.client.get(endpoint, **params)
                response = data.get("response") or []
                if response:
                    return response
            except RuntimeError:
                continue
        return []

    def update_player_row(self, row: int, profile: dict[str, Any], clubs: list[tuple[str, int | None, int | None]], note: str) -> None:
        self.set_cell(self.players_ws, row, "birth_year", profile.get("birth_year"), self.player_headers)
        self.set_cell(self.players_ws, row, "country", profile.get("country"), self.player_headers)
        if profile.get("photo_url"):
            self.set_cell(self.players_ws, row, "photo_url", profile.get("photo_url"), self.player_headers)
        if clubs:
            clubs_str = compact_clubs(clubs)
            self.set_cell(self.players_ws, row, "current_db_clubs", clubs_str, self.player_headers)
            if "wikidata_clubs" in self.player_headers:
                self.set_cell(self.players_ws, row, "wikidata_clubs", clubs_str, self.player_headers)
        self.set_cell(self.players_ws, row, "status", "api-football", self.player_headers)
        self.set_cell(self.players_ws, row, "notes", note, self.player_headers)

    def replace_club_rows(self, player_id: str, player_name: str, clubs: list[tuple[str, int | None, int | None]]) -> None:
        rows = self.transfer_rows_by_player.get(player_id, [])
        for row in reversed(rows):
            self.clubs_ws.delete_rows(row, 1)
        self.transfer_rows_by_player[player_id] = []
        for club_name, frm, to in clubs:
            self.clubs_ws.append([player_id, player_name, club_name, frm, to, "API-Football", "api", "Importado con API-Football"])

    def append_titles(self, player_id: str, player_name: str, trophies: list[dict[str, Any]]) -> int:
        count = 0
        existing = {
            (
                str(self.titles_ws.cell(row=r, column=self.title_headers["player_id"]).value or ""),
                str(self.titles_ws.cell(row=r, column=self.title_headers["title_name"]).value or ""),
                str(self.titles_ws.cell(row=r, column=self.title_headers["club_name"]).value or ""),
                str(self.titles_ws.cell(row=r, column=self.title_headers["year"]).value or ""),
            )
            for r in range(2, self.titles_ws.max_row + 1)
        }
        for trophy in trophies:
            league = trophy.get("league") or trophy.get("name")
            country = trophy.get("country")
            place = trophy.get("place")
            season = trophy.get("season")
            key = (player_id, str(league or ""), str(place or ""), str(season or ""))
            if key in existing:
                continue
            title_id = f"API-{player_id}-{self.titles_ws.max_row}"
            self.titles_ws.append(
                [
                    title_id,
                    player_id,
                    player_name,
                    league,
                    "official",
                    place,
                    parse_year(season) or season,
                    season,
                    country,
                    "API-Football",
                    "api",
                    "Importado con API-Football",
                ]
            )
            existing.add(key)
            count += 1
        return count

    def write_api_review_headers(self) -> None:
        if self.api_review_ws.max_row == 1 and self.api_review_ws.max_column == 1 and self.api_review_ws["A1"].value is None:
            self.api_review_ws.append(
                [
                    "player_id",
                    "name",
                    "api_status",
                    "issue",
                    "matched_name",
                    "api_player_id",
                    "clubs_found",
                    "titles_found",
                    "notes",
                ]
            )
        elif self.api_review_ws.cell(row=1, column=1).value != "player_id":
            self.api_review_ws.delete_rows(1, self.api_review_ws.max_row)
            self.api_review_ws.append(
                [
                    "player_id",
                    "name",
                    "api_status",
                    "issue",
                    "matched_name",
                    "api_player_id",
                    "clubs_found",
                    "titles_found",
                    "notes",
                ]
            )

    def append_review(self, player_id: str, name: str, api_status: str, issue: str, matched_name: str = "", api_player_id: Any = "", clubs_found: int = 0, titles_found: int = 0, notes: str = "") -> None:
        self.api_review_ws.append([player_id, name, api_status, issue, matched_name, api_player_id, clubs_found, titles_found, notes])

    def run(self, limit: int | None = None) -> None:
        self.write_api_review_headers()
        processed = 0
        for row in range(2, self.players_ws.max_row + 1):
            player_id = str(self.players_ws.cell(row=row, column=self.player_headers["player_id"]).value or "").strip()
            name = str(self.players_ws.cell(row=row, column=self.player_headers["name"]).value or "").strip()
            if not player_id or not name:
                continue
            if limit is not None and processed >= limit:
                break
            processed += 1
            try:
                match, error = self.find_best_player_match(name)
                if not match:
                    self.summary["no_match"] += 1
                    self.append_review(player_id, name, "no_match", error or "Sin match")
                    continue
                profile = self.extract_player_profile(match)
                api_player_id = profile.get("api_player_id")
                clubs = self.fetch_transfers(api_player_id) if api_player_id else []
                trophies = self.fetch_trophies(api_player_id) if api_player_id else []
                note = f"Actualizado con API-Football el {time.strftime('%Y-%m-%d')}."
                self.update_player_row(row, profile, clubs, note)
                if clubs:
                    self.replace_club_rows(player_id, name, clubs)
                titles_added = self.append_titles(player_id, name, trophies)
                self.append_review(
                    player_id,
                    name,
                    "ok" if clubs or trophies else "partial",
                    "" if (clubs or trophies) else "Sin clubs ni titulos en API",
                    matched_name=profile.get("full_name") or "",
                    api_player_id=api_player_id or "",
                    clubs_found=len(clubs),
                    titles_found=titles_added,
                    notes=note,
                )
                self.summary["updated"] += 1
                if not clubs:
                    self.summary["missing_clubs"] += 1
                if not trophies:
                    self.summary["missing_trophies"] += 1
            except Exception as exc:  # noqa: BLE001
                self.summary["errors"] += 1
                self.append_review(player_id, name, "error", str(exc))

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.output_path)

    def print_summary(self) -> None:
        print(json.dumps(dict(self.summary), ensure_ascii=False, indent=2))


def main() -> int:
    parser = argparse.ArgumentParser(description="Completa FutBrain con API-Football.")
    parser.add_argument("--api-key", required=True, help="API key de API-Football.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="Excel de entrada.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Excel de salida.")
    parser.add_argument("--limit", type=int, default=None, help="Limita la cantidad de jugadores para una prueba.")
    args = parser.parse_args()

    enricher = FutbrainApiEnricher(
        workbook_path=Path(args.input),
        output_path=Path(args.output),
        api_key=args.api_key,
    )
    enricher.run(limit=args.limit)
    enricher.print_summary()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
