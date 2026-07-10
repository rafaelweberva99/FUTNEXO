from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = ROOT / "futbrain_db_con_titulos.xlsx"
OUTPUT_JS = ROOT / "futbrain-db.js"
CURRENT_YEAR = 2026

TEXT_FIXES = {
    "S?o Paulo": "São Paulo",
    "SÃ£o Paulo": "São Paulo",
    "Am?rica Mineiro": "América Mineiro",
    "Be?ikta?": "Beşiktaş",
    "Inter Zapre?i?": "Inter Zaprešić",
    "Vit?ria de Guimar?es": "Vitória de Guimarães",
    "Uni?o S?o Jo?o": "União São João",
    "Atl?tico Mineiro": "Atlético Mineiro",
    "Fenerbah?e": "Fenerbahçe",
    "Claude Mak?l?l?": "Claude Makélélé",
    "Juan Sebasti?n Ver?n": "Juan Sebastián Verón",
    "Ra?l Gonz?lez": "Raúl González",
    "Robert Pir?s": "Robert Pirès",
    "Ronaldo Naz?rio": "Ronaldo Nazário",
    "MÃ¡s": "Más",
    "espa?ol": "español",
}


def normalize_text(value: str) -> str:
    text = str(repair_text(value) or "").strip().lower()
    replacements = {
        "á": "a",
        "à": "a",
        "ä": "a",
        "â": "a",
        "ã": "a",
        "é": "e",
        "è": "e",
        "ë": "e",
        "ê": "e",
        "í": "i",
        "ì": "i",
        "ï": "i",
        "î": "i",
        "ó": "o",
        "ò": "o",
        "ö": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ù": "u",
        "ü": "u",
        "û": "u",
        "ñ": "n",
        "ç": "c",
        "'": " ",
        "’": " ",
        ".": " ",
        "-": " ",
        "&": " and ",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    tokens = [token for token in text.split() if token not in {"fc", "f", "c", "club", "football", "futbol", "futebol", "cf", "ac", "as", "sc", "cd", "ca"}]
    return " ".join(tokens)


def canonical_club_name(value: str) -> str:
    value = repair_text(value)
    norm = normalize_text(value)
    if "brighton" in norm and "albion" in norm:
        return "Brighton"
    if norm.startswith("liverpool"):
        return "Liverpool"
    if norm.startswith("manchester united"):
        return "Manchester United"
    if norm.startswith("manchester city"):
        return "Manchester City"
    aliases = [
        ("Barcelona", ["barcelona"]),
        ("Real Madrid", ["real madrid", "real madrid club de futbol"]),
        ("Atlético Madrid", ["atletico madrid"]),
        ("Southampton", ["southampton", "southampton fc"]),
        ("Bayern Munich", ["bayern munich", "bayern munchen", "fc bayern munich", "fc bayern munchen"]),
        ("Borussia Dortmund", ["borussia dortmund"]),
        ("Inter Milan", ["inter milan", "internazionale", "fc internazionale milano"]),
        ("Milan", ["milan", "ac milan"]),
        ("Juventus", ["juventus", "juventus fc"]),
        ("PSG", ["psg", "paris saint germain"]),
        ("Lyon", ["lyon", "olympique lyonnais"]),
        ("Manchester United", ["manchester united", "manchester united fc"]),
        ("Manchester City", ["manchester city", "manchester city fc"]),
        ("Chelsea", ["chelsea", "chelsea fc"]),
        ("Liverpool", ["liverpool", "liverpool fc"]),
        ("Arsenal", ["arsenal", "arsenal fc"]),
        ("Tottenham", ["tottenham", "tottenham hotspur", "tottenham hotspur fc"]),
        ("Sevilla", ["sevilla", "sevilla fc"]),
        ("Valencia", ["valencia", "valencia cf"]),
        ("Roma", ["roma", "as roma"]),
        ("Lazio", ["lazio", "ss lazio"]),
        ("Napoli", ["napoli", "ssc napoli"]),
        ("Porto", ["porto", "fc porto"]),
        ("Benfica", ["benfica", "sl benfica"]),
        ("Ajax", ["ajax", "afc ajax"]),
        ("Monaco", ["monaco", "as monaco"]),
        ("Sporting CP", ["sporting cp", "sporting clube de portugal"]),
        ("LAFC", ["los angeles fc", "los angeles football club", "lafc"]),
        ("LA Galaxy", ["la galaxy", "los angeles galaxy"]),
        ("Al Nassr", ["al nassr", "al nassr fc"]),
        ("Al Hilal", ["al hilal", "al hilal sfc"]),
        ("Al Ittihad", ["al ittihad", "al ittihad fc"]),
        ("Al Sadd", ["al sadd", "al sadd sc"]),
        ("Athletico Paranaense", ["athletico paranaense", "club athletico paranaense"]),
        ("Flamengo", ["flamengo", "clube de regatas do flamengo"]),
        ("Colo-Colo", ["colo colo", "club social y deportivo colo colo"]),
        ("Universidad de Chile", ["universidad de chile", "club universidad de chile"]),
        ("Boca Juniors", ["boca juniors", "club atletico boca juniors"]),
        ("River Plate", ["river plate", "club atletico river plate"]),
        ("Santos", ["santos", "santos fc"]),
        ("Palmeiras", ["palmeiras", "se palmeiras"]),
        ("São Paulo", ["sao paulo", "sao paulo fc"]),
        ("Fluminense", ["fluminense", "fluminense fc"]),
        ("Parma", ["parma", "parma calcio 1913"]),
        ("Inter Miami", ["inter miami", "club internacional de futbol miami"]),
        ("Newell's Old Boys", ["newell s old boys", "club atletico newell s old boys"]),
        ("Brighton", ["brighton hove albion", "brighton and hove albion", "brighton hove albion fc"]),
        ("West Ham United", ["west ham united", "west ham united fc"]),
        ("Leicester City", ["leicester city", "leicester city fc"]),
        ("Aston Villa", ["aston villa", "aston villa fc"]),
        ("Everton", ["everton", "everton fc"]),
        ("Wolverhampton", ["wolverhampton", "wolverhampton wanderers", "wolverhampton wanderers fc"]),
        ("Newcastle United", ["newcastle united", "newcastle united fc"]),
        ("Leeds United", ["leeds united", "leeds united fc"]),
        ("Fulham", ["fulham", "fulham fc"]),
        ("Galatasaray", ["galatasaray", "galatasaray sk"]),
        ("Fenerbahce", ["fenerbahce", "fenerbahce sk"]),
        ("Schalke 04", ["schalke 04", "fc schalke 04"]),
        ("Bayer Leverkusen", ["bayer leverkusen", "bayer 04 leverkusen"]),
        ("Wolfsburg", ["wolfsburg", "vfl wolfsburg"]),
        ("Red Bull Salzburg", ["red bull salzburg", "fc red bull salzburg"]),
        ("Marseille", ["marseille", "olympique de marseille"]),
        ("Lille", ["lille", "losc lille"]),
        ("Argentinos Juniors", ["argentinos juniors", "asociacion atletica argentinos juniors"]),
        ("Racing Club", ["racing club", "racing club de avellaneda"]),
        ("Corinthians", ["corinthians", "sc corinthians paulista"]),
        ("Vissel Kobe", ["vissel kobe"]),
        ("CF Montreal", ["cf montreal", "club de foot montreal"]),
        ("Odisha FC", ["odisha fc"]),
        ("Miami United", ["miami united", "miami united fc"]),
    ]
    for canonical, variants in aliases:
        if norm in variants:
            return canonical
    return value.strip()


def repair_text(value):
    if value is None:
        return None
    text = str(value)
    for bad, good in TEXT_FIXES.items():
        text = text.replace(bad, good)
    return text


def merge_club_spans(rows: list[dict]) -> list[dict]:
    ordered = sorted(
        rows,
        key=lambda item: (
            item.get("from") if item.get("from") is not None else 9999,
            item.get("to") if item.get("to") is not None else 9999,
            item["club"],
        ),
    )
    merged: list[dict] = []
    for row in ordered:
        if not merged:
            merged.append(row.copy())
            continue
        prev = merged[-1]
        prev_to = prev.get("to")
        row_from = row.get("from")
        same_club = prev["club"] == row["club"]
        contiguous = (
            prev_to is not None
            and row_from is not None
            and row_from <= prev_to + 1
        )
        if same_club and contiguous:
            next_to = row.get("to")
            if prev_to is None or next_to is None:
                prev["to"] = None
            else:
                prev["to"] = max(prev_to, next_to)
            continue
        merged.append(row.copy())
    return merged


def rows_as_dicts(ws):
    headers = [str(cell or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for values in ws.iter_rows(min_row=2, values_only=True):
        yield {headers[idx]: value for idx, value in enumerate(values)}


def main() -> None:
    wb = load_workbook(SOURCE_XLSX, read_only=True)
    ws_players = wb["Jugadores"]
    ws_clubs = wb["Jugador_Clubes"]

    players_meta: dict[str, dict] = {}
    player_order: list[str] = []

    for row in rows_as_dicts(ws_players):
        player_id = row.get("player_id")
        name = row.get("name")
        name = repair_text(name)
        if not player_id or not name:
            continue
        if player_id in players_meta:
            continue
        player_order.append(player_id)
        wiki_title = row.get("wiki_title") or row.get("wiki_title".replace("_", " ")) or name
        if not row.get("wiki_title") and row.get("wikipedia_url"):
            wiki_title = str(row["wikipedia_url"]).rstrip("/").split("/")[-1].replace("_", " ")
        wiki_title = repair_text(wiki_title)
        players_meta[player_id] = {
            "id": player_id,
            "name": name,
            "flag": row.get("flag") or "",
            "country": repair_text(row.get("country")) or None,
            "photo": row.get("photo_url") or row.get("photo") or None,
            "birthYear": int(row["birth_year"]) if row.get("birth_year") else None,
            "wiki": wiki_title or name,
            "wikipediaUrl": row.get("wikipedia_url") or None,
            "wikidataUrl": row.get("wikidata_url") or None,
            "status": repair_text(row.get("status")) or None,
            "notes": repair_text(row.get("notes")) or None,
            "clubs": [],
        }

    clubs_by_player: dict[str, list[dict]] = defaultdict(list)
    for row in rows_as_dicts(ws_clubs):
        player_id = row.get("player_id")
        club_name = repair_text(row.get("club_name"))
        from_year = row.get("from_year")
        to_year = row.get("to_year")
        if not player_id or not club_name or player_id not in players_meta:
            continue
        clubs_by_player[player_id].append(
            {
                "club": canonical_club_name(str(club_name)),
                "from": int(from_year) if from_year else None,
                "to": int(to_year) if to_year else None,
            }
        )

    players: list[dict] = []
    seen_names: set[str] = set()
    legacy: dict[str, dict] = {}

    for player_id in player_order:
        meta = players_meta[player_id]
        if meta["name"] in seen_names:
            continue
        seen_names.add(meta["name"])
        merged_clubs = merge_club_spans(clubs_by_player.get(player_id, []))
        player = {
            "id": meta["id"],
            "name": meta["name"],
            "flag": meta["flag"],
            "country": meta["country"],
            "photo": meta["photo"],
            "birthYear": meta["birthYear"],
            "wiki": meta["wiki"],
            "wikipediaUrl": meta["wikipediaUrl"],
            "wikidataUrl": meta["wikidataUrl"],
            "status": meta["status"],
            "notes": meta["notes"],
            "clubs": merged_clubs,
        }
        players.append(player)
        legacy[player["name"]] = {
            "flag": player["flag"],
            "wiki": player["wiki"],
            "photo": player["photo"],
            "birthYear": player["birthYear"],
            "country": player["country"],
            "clubs": [
                {
                    "n": club["club"],
                    "f": club["from"],
                    "t": club["to"] if club["to"] is not None else CURRENT_YEAR,
                }
                for club in merged_clubs
            ],
        }

    content = f"""// Auto-generated from {SOURCE_XLSX.name}
(function(){{
  const FUTBRAIN_DB = {json.dumps(players, ensure_ascii=False, indent=2)};

  function normalizeText(value){{
    return String(value || '')
      .toLowerCase()
      .normalize('NFD')
      .replace(/[\\u0300-\\u036f]/g, '')
      .trim();
  }}

  function getPlayerById(id){{
    return FUTBRAIN_DB.find((player) => player.id === id) || null;
  }}

  function getPlayersByClub(clubName){{
    const needle = normalizeText(clubName);
    return FUTBRAIN_DB.filter((player) => player.clubs.some((club) => normalizeText(club.club) === needle));
  }}

  function getPlayersByCountry(country){{
    const needle = normalizeText(country);
    return FUTBRAIN_DB.filter((player) => normalizeText(player.country) === needle);
  }}

  function searchPlayers(query){{
    const needle = normalizeText(query);
    if (!needle) return [];
    return FUTBRAIN_DB.filter((player) => normalizeText(player.name).includes(needle));
  }}

  const FUTBRAIN_DB_LEGACY = {json.dumps(legacy, ensure_ascii=False, indent=2)};

  window.FUTBRAIN_DB = FUTBRAIN_DB;
  window.FUTBRAIN_DB_LEGACY = FUTBRAIN_DB_LEGACY;
  window.FUTBRAIN_DB_HELPERS = {{
    getPlayerById,
    getPlayersByClub,
    getPlayersByCountry,
    searchPlayers,
  }};
}})();
"""

    OUTPUT_JS.write_text(content, encoding="utf-8")
    total_clubs = sum(len(player["clubs"]) for player in players)
    active_clubs = sum(1 for player in players for club in player["clubs"] if club["to"] is None)
    print(f"Generado: {OUTPUT_JS}")
    print(f"Jugadores: {len(players)}")
    print(f"Registros jugador-club: {total_clubs}")
    print(f"Clubes activos con to=null: {active_clubs}")


if __name__ == "__main__":
    main()
