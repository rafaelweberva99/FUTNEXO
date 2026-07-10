from __future__ import annotations

import json
import subprocess
import time
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote
from urllib.request import urlopen

from openpyxl import Workbook
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parent
INDEX_PATH = ROOT / "index.html"
OUTPUT_PATH = ROOT / "futbrain_db_enriched.xlsx"
NODE_BIN = r"C:\Users\rafael.weber\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\bin\node.exe"


def json_get(url: str) -> dict:
    with urlopen(url) as response:
        return json.loads(response.read().decode("utf-8"))


def extract_db(index_path: Path) -> dict:
    node_script = f"""
const fs = require("fs");
const vm = require("vm");
const text = fs.readFileSync({json.dumps(str(index_path))}, "utf8");
const start = text.indexOf("const DB = {{");
const end = text.indexOf("const TRIVIA=[");
if (start === -1 || end === -1) {{
  throw new Error("No se pudo encontrar el bloque DB en index.html");
}}
const script = text.slice(start, end) + "\\n;globalThis.__DB__ = DB;";
const context = {{ console, globalThis: {{}} }};
vm.createContext(context);
vm.runInContext(script, context);
process.stdout.write(JSON.stringify(context.globalThis.__DB__));
"""
    temp_script = ROOT / "_extract_db_temp.js"
    temp_script.write_text(node_script, encoding="utf-8")
    try:
        result = subprocess.run(
            [NODE_BIN, str(temp_script)],
            cwd=str(ROOT),
            capture_output=True,
            check=True,
        )
        return json.loads(result.stdout.decode("utf-8"))
    finally:
        if temp_script.exists():
            temp_script.unlink()


def year_from_wikidata_time(value: str | None) -> int | None:
    if not value:
        return None
    try:
        return int(value[1:5])
    except Exception:
        return None


def wikipedia_lookup(title: str) -> dict:
    url = (
        "https://en.wikipedia.org/w/api.php?action=query"
        f"&titles={quote(title)}"
        "&prop=pageimages|pageprops"
        "&format=json"
        "&pithumbsize=400"
        "&origin=*"
    )
    data = json_get(url)
    pages = data.get("query", {}).get("pages", {})
    page = next(iter(pages.values()), {})
    return {
        "thumbnail": page.get("thumbnail", {}).get("source", ""),
        "wikibase_item": page.get("pageprops", {}).get("wikibase_item", ""),
    }


def chunked(values: list[str], size: int) -> list[list[str]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def wikidata_entities(ids: list[str]) -> dict:
    entities: dict = {}
    for chunk in chunked(ids, 40):
        url = (
            "https://www.wikidata.org/w/api.php?action=wbgetentities"
            f"&ids={'|'.join(chunk)}&languages=en&format=json&origin=*"
        )
        data = json_get(url)
        entities.update(data.get("entities", {}))
        time.sleep(0.05)
    return entities


def get_label(entity: dict) -> str:
    return entity.get("labels", {}).get("en", {}).get("value", "")


def simplify_current_clubs(clubs: list[dict]) -> list[tuple[str, int | None, int | None]]:
    return [(club.get("n", ""), club.get("f"), club.get("t")) for club in clubs]


def auto_fit(ws) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        longest = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            longest = max(longest, len(value))
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 70)


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def build_enriched_workbook(base_db: dict) -> Workbook:
    player_names = sorted(base_db.keys())
    player_meta: dict[str, dict] = {}
    wikidata_ids: list[str] = []

    for name in player_names:
        wiki_title = base_db[name].get("wiki", name)
        lookup = wikipedia_lookup(wiki_title)
        player_meta[name] = {
            "wiki_title": wiki_title,
            "wikipedia_url": f"https://en.wikipedia.org/wiki/{wiki_title.replace(' ', '_')}",
            "photo_url": lookup["thumbnail"],
            "wikidata_id": lookup["wikibase_item"],
        }
        if lookup["wikibase_item"]:
            wikidata_ids.append(lookup["wikibase_item"])
        time.sleep(0.05)

    player_entities = wikidata_entities(sorted(set(wikidata_ids)))

    needed_aux_ids: set[str] = set()
    for entity in player_entities.values():
        claims = entity.get("claims", {})
        for claim in claims.get("P54", []):
            mainsnak = claim.get("mainsnak", {})
            datavalue = mainsnak.get("datavalue", {})
            club_id = datavalue.get("value", {}).get("id")
            if club_id:
                needed_aux_ids.add(club_id)
        for claim in claims.get("P27", []):
            mainsnak = claim.get("mainsnak", {})
            datavalue = mainsnak.get("datavalue", {})
            country_id = datavalue.get("value", {}).get("id")
            if country_id:
                needed_aux_ids.add(country_id)

    aux_entities = wikidata_entities(sorted(needed_aux_ids))

    wb = Workbook()
    ws_players = wb.active
    ws_players.title = "Jugadores"
    ws_players.append(
        [
            "player_id",
            "name",
            "flag",
            "wiki_title",
            "wikipedia_url",
            "wikidata_id",
            "photo_url",
            "birth_year",
            "country",
            "current_db_clubs",
            "wikidata_club_rows",
            "status",
            "notes",
        ]
    )

    ws_player_clubs = wb.create_sheet("Jugador_Clubes")
    ws_player_clubs.append(
        [
            "player_id",
            "player_name",
            "club_name",
            "from_year",
            "to_year",
            "source",
            "verification",
            "notes",
        ]
    )

    club_to_players: dict[str, set[str]] = defaultdict(set)
    review_rows: list[list] = []

    for idx, name in enumerate(player_names, start=1):
        player_id = f"P{idx:04d}"
        data = base_db[name]
        meta = player_meta[name]
        entity = player_entities.get(meta["wikidata_id"], {})
        claims = entity.get("claims", {})

        birth_year = None
        if claims.get("P569"):
            birth_year = year_from_wikidata_time(
                claims["P569"][0].get("mainsnak", {}).get("datavalue", {}).get("value", {}).get("time")
            )

        country = ""
        if claims.get("P27"):
            country_id = (
                claims["P27"][0].get("mainsnak", {}).get("datavalue", {}).get("value", {}).get("id")
            )
            if country_id:
                country = get_label(aux_entities.get(country_id, {}))

        wikidata_rows = []
        for claim in claims.get("P54", []):
            if claim.get("rank") == "deprecated":
                continue
            mainsnak = claim.get("mainsnak", {})
            datavalue = mainsnak.get("datavalue", {})
            club_id = datavalue.get("value", {}).get("id")
            if not club_id:
                continue
            qualifiers = claim.get("qualifiers", {})
            from_year = None
            to_year = None
            if qualifiers.get("P580"):
                from_year = year_from_wikidata_time(
                    qualifiers["P580"][0].get("datavalue", {}).get("value", {}).get("time")
                )
            if qualifiers.get("P582"):
                to_year = year_from_wikidata_time(
                    qualifiers["P582"][0].get("datavalue", {}).get("value", {}).get("time")
                )
            club_name = get_label(aux_entities.get(club_id, {})) or club_id
            wikidata_rows.append((club_name, from_year, to_year))

        wikidata_rows.sort(key=lambda item: (item[1] is None, item[1] or 9999, item[2] or 9999, item[0]))
        for club_name, from_year, to_year in wikidata_rows:
            club_to_players[club_name].add(name)
            ws_player_clubs.append(
                [player_id, name, club_name, from_year, to_year, "Wikidata/Wikipedia", "", ""]
            )

        current_rows = simplify_current_clubs(data.get("clubs", []))
        current_comp = [f"{club} ({f or '?'}-{t or '?'})" for club, f, t in current_rows]
        wikidata_comp = [f"{club} ({f or '?'}-{t or '?'})" for club, f, t in wikidata_rows]

        status = "ok"
        notes = ""
        if not meta["wikidata_id"]:
            status = "sin_wikidata"
            notes = "No se encontró item de Wikidata desde Wikipedia."
        elif not wikidata_rows:
            status = "sin_clubes"
            notes = "Wikidata no devolvió historial de clubes."
        elif current_comp != wikidata_comp:
            status = "revisar"
            notes = "La trayectoria actual y la de Wikidata no coinciden exactamente."

        ws_players.append(
            [
                player_id,
                name,
                data.get("flag", ""),
                meta["wiki_title"],
                meta["wikipedia_url"],
                meta["wikidata_id"],
                meta["photo_url"],
                birth_year,
                country,
                " | ".join(current_comp),
                " | ".join(wikidata_comp),
                status,
                notes,
            ]
        )

        if status != "ok":
            review_rows.append(
                [
                    player_id,
                    name,
                    status,
                    " | ".join(current_comp),
                    " | ".join(wikidata_comp),
                    meta["wikipedia_url"],
                    notes,
                ]
            )

    style_header(ws_players)
    auto_fit(ws_players)
    style_header(ws_player_clubs)
    auto_fit(ws_player_clubs)

    ws_clubs = wb.create_sheet("Clubes")
    ws_clubs.append(
        ["club_id", "club_name", "player_count", "players", "country", "league", "badge_url", "notes"]
    )
    for idx, club_name in enumerate(sorted(club_to_players.keys()), start=1):
        players = sorted(club_to_players[club_name])
        ws_clubs.append([f"C{idx:04d}", club_name, len(players), " | ".join(players), "", "", "", ""])
    style_header(ws_clubs)
    auto_fit(ws_clubs)

    ws_coaches = wb.create_sheet("Entrenadores")
    ws_coaches.append(["coach_id", "name", "country", "photo_url", "teams_coached", "notes"])
    style_header(ws_coaches)
    auto_fit(ws_coaches)

    ws_review = wb.create_sheet("Revisar")
    ws_review.append(
        ["player_id", "name", "status", "current_db_clubs", "wikidata_clubs", "wikipedia_url", "notes"]
    )
    for row in review_rows:
        ws_review.append(row)
    style_header(ws_review)
    auto_fit(ws_review)

    ws_summary = wb.create_sheet("Resumen")
    ws_summary.append(["Métrica", "Valor"])
    ws_summary.append(["Jugadores cargados", len(player_names)])
    ws_summary.append(["Clubes únicos detectados", len(club_to_players)])
    ws_summary.append(["Entrenadores cargados", 0])
    ws_summary.append(["Jugadores OK", sum(1 for row in ws_players.iter_rows(min_row=2, values_only=True) if row[11] == "ok")])
    ws_summary.append(["Jugadores para revisar", len(review_rows)])
    ws_summary.append(["Jugadores sin foto_url", '=COUNTBLANK(Jugadores!G2:G9999)'])
    ws_summary.append(["Jugadores sin Wikidata", '=COUNTIF(Jugadores!L2:L9999,"sin_wikidata")'])
    ws_summary.append(["Jugadores sin clubes en Wikidata", '=COUNTIF(Jugadores!L2:L9999,"sin_clubes")'])
    ws_summary.append(["Nota", "Los casos no resueltos o diferentes quedaron en la hoja Revisar."])
    style_header(ws_summary)
    auto_fit(ws_summary)

    return wb


def main() -> None:
    base_db = extract_db(INDEX_PATH)
    workbook = build_enriched_workbook(base_db)
    workbook.save(OUTPUT_PATH)
    print(f"Excel enriquecido: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
