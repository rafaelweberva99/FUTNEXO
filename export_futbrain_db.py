from __future__ import annotations

import json
import subprocess
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parent
INDEX_PATH = ROOT / "index.html"
OUTPUT_PATH = ROOT / "futbrain_db.xlsx"
NODE_FALLBACK = Path(r"C:\Users\rafael.weber\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\bin\node.exe")


def extract_db(index_path: Path) -> dict:
    node_script = f"""
const fs = require("fs");
const vm = require("vm");
const text = fs.readFileSync({json.dumps(str(index_path))}, "utf8");
const start = text.indexOf("const DB = {{");
const end = text.indexOf("const TRIVIA=[");
if (start === -1 || end === -1) {{
  throw new Error("No se pudo encontrar el bloque DB en index.html");
}}
const script = text.slice(start, end) + "\\n;globalThis.__DB__ = DB;";
const context = {{ console, globalThis: {{}} }};
vm.createContext(context);
vm.runInContext(script, context);
process.stdout.write(JSON.stringify(context.globalThis.__DB__));
"""
    temp_script = ROOT / "_extract_db_temp.js"
    temp_script.write_text(node_script, encoding="utf-8")
    node_bin = str(NODE_FALLBACK)
    try:
        result = subprocess.run(
            [node_bin, str(temp_script)],
            cwd=str(ROOT),
            capture_output=True,
            check=True,
        )
        return json.loads(result.stdout.decode("utf-8"))
    finally:
        if temp_script.exists():
            temp_script.unlink()


def auto_fit(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[letter].width = min(max(length + 2, 12), 60)


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def build_workbook(db: dict) -> Workbook:
    wb = Workbook()
    players_ws = wb.active
    players_ws.title = "Jugadores"

    clubs_to_players: dict[str, set[str]] = defaultdict(set)
    club_rows: list[list] = []

    player_headers = [
        "player_id",
        "name",
        "flag",
        "wiki_title",
        "wikipedia_url",
        "photo_url",
        "photo_status",
        "club_count",
        "clubs_played",
        "notes",
    ]
    players_ws.append(player_headers)

    for idx, name in enumerate(sorted(db.keys()), start=1):
        data = db[name]
        wiki_title = data.get("wiki", name)
        clubs = data.get("clubs", [])
        club_names = []
        for club in clubs:
            club_name = club.get("n", "").strip()
            if not club_name:
                continue
            club_names.append(club_name)
            clubs_to_players[club_name].add(name)
            club_rows.append(
                [
                    f"P{idx:04d}",
                    name,
                    club_name,
                    club.get("f"),
                    club.get("t"),
                    f"{club.get('f', '')}-{club.get('t', '')}",
                    "index.html",
                    "",
                ]
            )

        wikipedia_url = f"https://en.wikipedia.org/wiki/{wiki_title.replace(' ', '_')}"
        players_ws.append(
            [
                f"P{idx:04d}",
                name,
                data.get("flag", ""),
                wiki_title,
                wikipedia_url,
                "",
                "pendiente",
                len(club_names),
                " | ".join(club_names),
                "",
            ]
        )

    style_header(players_ws)
    auto_fit(players_ws)

    player_clubs_ws = wb.create_sheet("Jugador_Clubes")
    player_clubs_ws.append(
        [
            "player_id",
            "player_name",
            "club_name",
            "from_year",
            "to_year",
            "period",
            "source",
            "verified",
        ]
    )
    for row in club_rows:
        player_clubs_ws.append(row)
    style_header(player_clubs_ws)
    auto_fit(player_clubs_ws)

    clubs_ws = wb.create_sheet("Clubes")
    clubs_ws.append(
        [
            "club_id",
            "club_name",
            "player_count",
            "players",
            "country",
            "league",
            "badge_url",
            "notes",
        ]
    )
    for idx, club_name in enumerate(sorted(clubs_to_players.keys()), start=1):
        players = sorted(clubs_to_players[club_name])
        clubs_ws.append(
            [
                f"C{idx:04d}",
                club_name,
                len(players),
                " | ".join(players),
                "",
                "",
                "",
                "",
            ]
        )
    style_header(clubs_ws)
    auto_fit(clubs_ws)

    coaches_ws = wb.create_sheet("Entrenadores")
    coaches_ws.append(
        [
            "coach_id",
            "name",
            "country",
            "photo_url",
            "teams_coached",
            "notes",
        ]
    )
    style_header(coaches_ws)
    auto_fit(coaches_ws)

    summary_ws = wb.create_sheet("Resumen")
    summary_ws.append(["Métrica", "Valor"])
    summary_ws.append(["Jugadores cargados", players_ws.max_row - 1])
    summary_ws.append(["Clubes únicos", clubs_ws.max_row - 1])
    summary_ws.append(["Entrenadores cargados", 0])
    summary_ws.append(["Relaciones jugador-club", player_clubs_ws.max_row - 1])
    summary_ws.append(["Jugadores sin foto_url", '=COUNTBLANK(Jugadores!F2:F9999)'])
    summary_ws.append(["Jugadores con wiki_title", '=COUNTA(Jugadores!D2:D9999)'])
    summary_ws.append(["Jugadores con 1 solo club", '=COUNTIF(Jugadores!H2:H9999,1)'])
    summary_ws.append(["Jugadores con 5+ clubes", '=COUNTIF(Jugadores!H2:H9999,\">=5\")'])
    summary_ws.append(["Observación", "La hoja Entrenadores está vacía porque hoy el juego no tiene esa base cargada."])
    style_header(summary_ws)
    auto_fit(summary_ws)

    return wb


def main() -> None:
    db = extract_db(INDEX_PATH)
    wb = build_workbook(db)
    wb.save(OUTPUT_PATH)
    print(f"Excel generado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
